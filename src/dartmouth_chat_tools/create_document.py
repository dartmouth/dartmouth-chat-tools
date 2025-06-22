import io
import logging
import tempfile
from typing import Literal

import pandas as pd
import pypandoc
from fastapi import UploadFile
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font
from pydantic import BaseModel

from open_webui.functions import UserModel
from open_webui.models.users import Users
from open_webui.routers.files import upload_file as owui_upload_file


log = logging.getLogger(__name__)


class Tools:
    class Valves(BaseModel):
        open_webui_url: str = "http://localhost:8080/"
        spreadsheet_header_background_color: str = "#00693E"
        spreadsheet_header_font_color: str = "#FFFFFF"
        spreadsheet_even_row_background_color: str = "#c4dd88"
        spreadsheet_even_row_font_color: str = "#000000"
        spreadsheet_odd_row_background_color: str = "#FFFFFF"
        spreadsheet_odd_row_font_color: str = "#000000"

    def __init__(self):
        self.valves = self.Valves()

    async def create_document(
        self,
        __user__: dict[str, str],
        content: str,
        output_format: Literal["docx", "pptx", "bibtex", "latex", "pdf"],
        filename_stem: str,
        input_format: Literal[
            "md", "html", "xml", "latex", "bibtex", "biblatex"
        ] = "md",
        template: Literal["standard", "dartmouth"] = "dartmouth",
    ) -> str:
        """Turns a string into a document of the specified format and returns the
        download URL.

        This uses pandoc under the hood, so format the content accordingly. For pptx,
        that means to include a title block at the top that needs to use the following syntax:

        ```markdown
        ---
        title: Presentation Title
        author: Author Name
        ---
        ```

        The filename_stem should not include the file extension.

        Do not use this tool unless the user specifically asks for a file to download.
        You must repeat the download URL to the user as a link.
        Remind the user to install [Dartmouth's Typefaces](https://communications.dartmouth.edu/guides-and-tools/design-guidelines/dartmouth-typefaces).
        You can also offer to alternatively generate the file with the standard plain template.
        """

        user = __user__

        # Create the filename (potential name collisions are handled by
        # OWUI's file upload)
        filename = f"{filename_stem}.{output_format}"

        # Create a temporary file
        with tempfile.NamedTemporaryFile(
            suffix=f".{output_format}", delete=True
        ) as temp_file:
            temp_path = temp_file.name

            extra_args = _get_extra_args(input_format, output_format, template)

            pypandoc.convert_text(
                source=content,
                to=output_format,
                format=input_format,
                outputfile=temp_path,
                extra_args=extra_args,
            )

            user_id = __user__.get("id")
            if user_id is None:
                raise Exception(f"No user ID available: {__user__}")

            user_obj = Users.get_user_by_id(user_id)
            if not user_obj:
                raise Exception(f"User not found with ID: {user_id}")

            content_type = _get_content_type(output_format)
            metadata = {
                "description": f"Document created with pandoc - {input_format} to {output_format}"
            }

            try:
                file_url = _upload_file(
                    user=user_obj,
                    filename=filename,
                    file_path=temp_path,
                    content_type=content_type,
                    metadata=metadata,
                    base_url=self.valves.open_webui_url,
                )

            except Exception as e:
                log.error(f"Failed to upload generated document: {str(e)}")
                raise Exception(f"Failed to upload generated document: {str(e)}")

            return file_url

    async def create_spreadsheet(
        self,
        __user__: dict,
        content: str,
        filename_stem: str,
        table_name: str = "DataTable",
        sheet_name: str = "Sheet1",
        output_format: Literal["xlsx"] = "xlsx",
        input_format: Literal["csv"] = "csv",
    ):
        """Create an Excel spreadsheet of the input data. Input data is expected in CSV format.
        The output Excel workbook will contain the data in a named table.
        Names for the table and the worksheet can be provided.

        The filename_stem should not include the file extension.

        Do not use this tool unless the user specifically asks for a file to download.
        You must repeat the download URL to the user as a link.
        """

        # Create the filename (potential name collisions are handled by
        # OWUI's file upload)
        filename = f"{filename_stem}.{output_format}"

        # Read CSV string into a pandas DataFrame
        csv_io = io.StringIO(content)
        df = pd.read_csv(csv_io)

        # Create a new workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Write DataFrame to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Define the table range (from A1 to last column and row with data)
        last_column = ws.max_column
        last_row = ws.max_row
        table_range = f"A1:{ws.cell(row=last_row, column=last_column).coordinate}"

        table = Table(displayName=table_name, ref=table_range)

        style = TableStyleInfo(
            name="TableStyleMedium9",  # Base style
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        table.tableStyleInfo = style

        # Add the table to the worksheet
        ws.add_table(table)

        # Apply styling
        header_color = self.valves.spreadsheet_header_background_color.lstrip("#")
        header_font_color = self.valves.spreadsheet_header_font_color.lstrip("#")
        even_row_color = self.valves.spreadsheet_even_row_background_color.lstrip("#")
        even_row_font_color = self.valves.spreadsheet_even_row_font_color.lstrip("#")
        odd_row_color = self.valves.spreadsheet_odd_row_background_color.lstrip("#")
        odd_row_font_color = self.valves.spreadsheet_odd_row_font_color.lstrip("#")

        header_fill = PatternFill(
            start_color=header_color, end_color=header_color, fill_type="solid"
        )
        header_font = Font(color=header_font_color, bold=True)

        even_row_fill = PatternFill(
            start_color=even_row_color, end_color=even_row_color, fill_type="solid"
        )
        even_row_font = Font(color=even_row_font_color, bold=False)

        odd_row_fill = PatternFill(
            start_color=odd_row_color, end_color=odd_row_color, fill_type="solid"
        )
        odd_row_font = Font(color=odd_row_font_color, bold=False)

        # Style the header row (row 1)
        for col in range(1, last_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font

        # Style alternating data rows (skip header row)
        for row in range(2, last_row + 1):
            if row % 2 == 0:
                for col in range(1, last_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = even_row_fill
                    cell.font = even_row_font
            if row % 2 != 0:
                for col in range(1, last_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = odd_row_fill
                    cell.font = odd_row_font

        # Save to temporary file and upload:
        with tempfile.NamedTemporaryFile(
            suffix=f".{output_format}", delete=True
        ) as temp_file:
            temp_path = temp_file.name

            # Save the workbook
            wb.save(temp_path)

            # Upload file
            user_id = __user__.get("id")
            if user_id is None:
                raise Exception(f"No user ID available: {__user__}")

            user_obj = Users.get_user_by_id(user_id)
            if not user_obj:
                raise Exception(f"User not found with ID: {user_id}")

            content_type = _get_content_type(output_format)
            metadata = {
                "description": f"Document created with openpyxl - {input_format} to {output_format}"
            }

            try:
                file_url = _upload_file(
                    user=user_obj,
                    filename=filename,
                    file_path=temp_path,
                    content_type=content_type,
                    metadata=metadata,
                    base_url=self.valves.open_webui_url,
                )

            except Exception as e:
                log.error(f"Failed to upload generated spreadsheet: {str(e)}")
                raise Exception(f"Failed to upload generated spreadsheet: {str(e)}")

            return file_url


class _MockRequest:
    """A minimal fake request object"""

    def __init__(self) -> None:
        from types import SimpleNamespace

        self.app = SimpleNamespace()
        self.app.state = SimpleNamespace()
        self.app.state.config = SimpleNamespace()
        self.app.state.config.ALLOWED_FILE_EXTENSIONS = None
        self.app.state.config.STT_SUPPORTED_CONTENT_TYPES = None
        self.app.state.config.CONTENT_EXTRACTION_ENGINE = None


def _create_mock_request() -> _MockRequest:
    """Create a minimal request object to use API endpoints directly"""

    request_obj = _MockRequest()

    return request_obj


def _get_content_type(output_format: str) -> str:
    """Get the content type based on the output format"""
    content_types = {
        "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        "pdf": "application/pdf",
    }
    return content_types.get(output_format, "application/octet-stream")


def _get_extra_args(
    input_format: str, output_format: str, template: Literal["standard", "dartmouth"]
) -> list[str]:
    """Get extra arguments for the pandoc invocation"""
    extra_args = []
    if template == "dartmouth":
        if output_format == "docx":
            extra_args.append(
                "--reference-doc=./data/resources/dartmouth-template.docx"
            )
        elif output_format == "pptx":
            extra_args.append(
                "--reference-doc=./data/resources/dartmouth-template.pptx"
            )
    return extra_args


def _upload_file(
    user: UserModel,
    filename: str,
    file_path: str,
    content_type: str,
    metadata: dict[str, str],
    base_url: str,
):
    """Upload the file using OWUI's file upload endpoint"""

    with open(file_path, "rb") as file_content:
        upload_file_obj = UploadFile(
            filename=filename,
            file=file_content,
        )
        # Set the content_type as a header
        # This is a workaround since we can't set the content_type attribute
        # directly
        upload_file_obj.headers = {"content-type": content_type}

        request_obj = _create_mock_request()

        # Call OWUI's upload_file function directly with a fake request
        file_response = owui_upload_file(
            request=request_obj,  # type: ignore
            file=upload_file_obj,
            metadata=metadata,
            process=False,
            internal=True,
            user=user,
        )

    file_id = file_response.id

    file_url = f"{base_url.strip('/')}/api/v1/files/{file_id}/content"
    return file_url
